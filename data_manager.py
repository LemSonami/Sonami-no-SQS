from __future__ import annotations

import json
import sys
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Dict, Iterable, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from crypto_utils import decrypt_bytes, encrypt_bytes

if getattr(sys, 'frozen', False):
    BASE_DIR = Path(sys.executable).resolve().parent
else:
    BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
USERS_XLSX_FILE = DATA_DIR / "users.xlsx"
USERS_SUDAUSERS_FILE = DATA_DIR / "users.sudausers"
QUESTIONS_SUDASQS_FILE = DATA_DIR / "questions.sudasqs"
SCORES_ILOVESUDA_FILE = DATA_DIR / "scores.ilovesuda"
SETTINGS_FILE = DATA_DIR / "settings.json"
LOGIN_CACHE_FILE = DATA_DIR / "login_cache.json"
AVATARS_FILE = DATA_DIR / "avatars.json"

DEFAULT_SETTINGS = {
    "font_family": "Microsoft YaHei",
    "api_key": "",
    "colors": {
        "BG": "#f4f7fb",
        "PANEL": "#ffffff",
        "TEXT": "#1f2937",
        "MUTED": "#64748b",
        "PRIMARY": "#2563eb",
        "SUCCESS": "#16a34a",
        "DANGER": "#dc2626",
        "BORDER": "#d8dee9",
    },
}

@dataclass
class Question:
    id: int
    qtype: str
    text: str
    options: List[str]
    answer: str
    analysis: str
    score: int = 10

    def to_row(self) -> List[object]:
        opts = self.options + ["", "", "", ""]
        return [
            self.id,
            self.qtype,
            self.text,
            opts[0],
            opts[1],
            opts[2],
            opts[3],
            self.answer,
            self.analysis,
            self.score,
        ]

class UserDB:

    def __init__(self) -> None:
        self._enc_path = USERS_SUDAUSERS_FILE
        self._xlsx_path = USERS_XLSX_FILE
        self._users: list[dict] = []

    def load_users(self) -> list[dict]:

        if self._enc_path.exists():
            return self._load_from_encrypted()

        if self._xlsx_path.exists():
            users = self._load_from_xlsx()
            self._encrypt_users_file()
            return users
        return []

    def _load_from_xlsx(self) -> list[dict]:
        try:
            wb = load_workbook(self._xlsx_path, read_only=True)
            sheet = wb.active
            users = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                student_id = str(row[0]).strip()
                name = str(row[1]).strip() if row[1] else ""
                password = str(row[2]).strip() if row[2] else ""
                role_code = str(row[3]).strip().upper() if row[3] else ""
                if role_code == "S":
                    role = "student"
                elif role_code == "T":
                    role = "teacher"
                else:
                    continue
                users.append({
                    "student_id": student_id,
                    "name": name,
                    "password": password,
                    "role": role,
                })
            wb.close()
            self._users = users
            return users
        except Exception as exc:
            raise RuntimeError(f"用户文件读取失败：{exc}") from exc

    def _load_from_encrypted(self) -> list[dict]:
        try:
            raw = decrypt_bytes(self._enc_path.read_bytes())
            wb = load_workbook(BytesIO(raw), read_only=True)
            sheet = wb.active
            users = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                student_id = str(row[0]).strip()
                name = str(row[1]).strip() if row[1] else ""
                password = str(row[2]).strip() if row[2] else ""
                role_code = str(row[3]).strip().upper() if row[3] else ""
                if role_code == "S":
                    role = "student"
                elif role_code == "T":
                    role = "teacher"
                else:
                    continue
                users.append({
                    "student_id": student_id,
                    "name": name,
                    "password": password,
                    "role": role,
                })
            wb.close()
            self._users = users
            return users
        except Exception as exc:
            raise RuntimeError(f"加密用户文件读取失败：{exc}") from exc

    def _encrypt_users_file(self) -> None:
        if not self._xlsx_path.exists():
            return
        try:
            raw = self._xlsx_path.read_bytes()
            self._enc_path.write_bytes(encrypt_bytes(raw))
        except Exception:
            pass

    def authenticate(self, student_id: str, name: str, password: str) -> Optional[dict]:
        sid = student_id.strip()
        nm = name.strip()
        pw = password.strip()
        for user in self._users:
            if user["student_id"] == sid and user["name"] == nm and user["password"] == pw:
                result = dict(user)
                result["username"] = user["student_id"]
                return result
        return None

    def is_registered(self, student_id: str) -> bool:
        sid = student_id.strip()
        return any(u["student_id"] == sid for u in self._users)

class QuestionBank:

    HEADERS = [
        "id",
        "type",
        "question",
        "option_a",
        "option_b",
        "option_c",
        "option_d",
        "answer",
        "analysis",
        "score",
    ]

    def __init__(self, path: Path = QUESTIONS_SUDASQS_FILE) -> None:
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self._migrate_from_xlsx()
        self._ensure_workbook()

    def _migrate_from_xlsx(self) -> None:
        old_path = self.path.parent / "questions.xlsx"
        if not old_path.exists() or self.path.exists():
            return
        try:
            data = old_path.read_bytes()
            self.path.write_bytes(encrypt_bytes(data))
            old_path.rename(old_path.with_suffix(".xlsx.bak"))
        except Exception:
            pass

    def _ensure_workbook(self) -> None:
        if self.path.exists():
            return
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "questions"
        sheet.append(self.HEADERS)
        self._style_header(sheet)
        self._save_workbook(workbook)

    def _save_workbook(self, workbook) -> None:
        buf = BytesIO()
        workbook.save(buf)
        self.path.write_bytes(encrypt_bytes(buf.getvalue()))

    def _style_header(self, sheet) -> None:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2F5597")

    def _load_sheet(self):
        try:
            raw = decrypt_bytes(self.path.read_bytes())
            workbook = load_workbook(BytesIO(raw))
            return workbook, workbook["questions"]
        except FileNotFoundError:
            self._ensure_workbook()
            raw = decrypt_bytes(self.path.read_bytes())
            workbook = load_workbook(BytesIO(raw))
            return workbook, workbook["questions"]
        except Exception as exc:
            raise RuntimeError(f"题库文件读取失败：{exc}") from exc

    def load_questions(self, qtype: Optional[str] = None) -> List[Question]:
        _, sheet = self._load_sheet()
        questions: List[Question] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            options = [str(value) for value in row[3:7] if value not in (None, "")]
            question = Question(
                id=int(row[0]),
                qtype=str(row[1]).strip(),
                text=str(row[2]).strip(),
                options=options,
                answer=str(row[7]).strip(),
                analysis=str(row[8] or "").strip(),
                score=int(row[9] or 10),
            )
            if qtype is None or question.qtype == qtype:
                questions.append(question)
        return questions

    def save_question(
        self,
        qtype: str,
        text: str,
        options: Iterable[str],
        answer: str,
        analysis: str,
        score: int = 10,
    ) -> Question:
        if qtype not in {"单选", "填空", "判断"}:
            raise ValueError("题型只能是：单选、填空、判断")
        if not text.strip():
            raise ValueError("题目不能为空")
        if not answer.strip():
            raise ValueError("答案不能为空")
        if score <= 0:
            raise ValueError("分值必须大于 0")

        workbook, sheet = self._load_sheet()
        next_id = self._next_id(sheet)
        clean_options = [item.strip() for item in options if item and item.strip()]
        question = Question(
            id=next_id,
            qtype=qtype,
            text=text.strip(),
            options=clean_options,
            answer=answer.strip(),
            analysis=analysis.strip(),
            score=score,
        )
        sheet.append(question.to_row())
        self._save_workbook(workbook)
        return question

    def delete_question(self, question_id: int) -> bool:
        workbook, sheet = self._load_sheet()
        for row_index in range(2, sheet.max_row + 1):
            if sheet.cell(row_index, 1).value == question_id:
                sheet.delete_rows(row_index, 1)
                self._save_workbook(workbook)
                return True
        return False

    def update_question(self, question: Question) -> bool:
        workbook, sheet = self._load_sheet()
        for row_index in range(2, sheet.max_row + 1):
            if sheet.cell(row_index, 1).value == question.id:
                for col_index, value in enumerate(question.to_row(), start=1):
                    sheet.cell(row_index, col_index, value)
                self._save_workbook(workbook)
                return True
        return False

    def _next_id(self, sheet) -> int:
        ids = [
            int(row[0])
            for row in sheet.iter_rows(min_row=2, values_only=True)
            if row and row[0] is not None
        ]
        return max(ids, default=0) + 1

    def reorder_questions(self, ordered_ids: List[int]) -> None:
        all_qs = {q.id: q for q in self.load_questions()}
        workbook, sheet = self._load_sheet()

        for _ in range(sheet.max_row, 1, -1):
            sheet.delete_rows(2)

        for new_id, old_id in enumerate(ordered_ids, start=1):
            q = all_qs[old_id]
            q.id = new_id
            sheet.append(q.to_row())
        self._save_workbook(workbook)

class ScoreManager:

    HEADERS = [
        "username",
        "student_id",
        "student_name",
        "score",
        "total",
        "exam_time",
        "wrong_ids",
        "type_stats",
        "wrong_answers",
    ]

    def __init__(self, path: Path = SCORES_ILOVESUDA_FILE) -> None:
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self._migrate_from_xlsx()
        self._ensure_workbook()

    def _migrate_from_xlsx(self) -> None:
        old_path = self.path.parent / "scores.xlsx"
        if not old_path.exists() or self.path.exists():
            return
        try:
            data = old_path.read_bytes()
            self.path.write_bytes(encrypt_bytes(data))
            old_path.rename(old_path.with_suffix(".xlsx.bak"))
        except Exception:
            pass

    def _ensure_workbook(self) -> None:
        if self.path.exists():
            return
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "scores"
        sheet.append(self.HEADERS)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="548235")
        self._save_workbook(workbook)

    def _save_workbook(self, workbook) -> None:
        buf = BytesIO()
        workbook.save(buf)
        self.path.write_bytes(encrypt_bytes(buf.getvalue()))

    def _load_sheet(self):
        try:
            raw = decrypt_bytes(self.path.read_bytes())
            workbook = load_workbook(BytesIO(raw))
            return workbook, workbook["scores"]
        except FileNotFoundError:
            self._ensure_workbook()
            raw = decrypt_bytes(self.path.read_bytes())
            workbook = load_workbook(BytesIO(raw))
            return workbook, workbook["scores"]
        except Exception as exc:
            raise RuntimeError(f"成绩文件读取失败：{exc}") from exc

    def record_score(
        self,
        user: Dict[str, str],
        score: int,
        total: int,
        wrong_ids: Iterable[int],
        type_stats: Dict[str, Dict[str, int]],
        wrong_answers: Optional[Dict[int, str]] = None,
    ) -> None:
        workbook, sheet = self._load_sheet()
        sheet.append(
            [
                user.get("username", ""),
                user.get("student_id", ""),
                user.get("name", ""),
                score,
                total,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                ",".join(str(item) for item in wrong_ids),
                json.dumps(type_stats, ensure_ascii=False),
                json.dumps({str(k): v for k, v in (wrong_answers or {}).items()}, ensure_ascii=False),
            ]
        )
        self._save_workbook(workbook)

    def import_record(self, record: Dict[str, object]) -> None:
        workbook, sheet = self._load_sheet()
        wrong_answers = record.get("wrong_answers", {})
        sheet.append(
            [
                str(record.get("username", "")),
                str(record.get("student_id", "")),
                str(record.get("student_name", "")),
                int(record.get("score", 0)),
                int(record.get("total", 0)),
                str(record.get("exam_time", "")),
                ",".join(str(i) for i in record.get("wrong_ids", [])),
                json.dumps(record.get("type_stats", {}), ensure_ascii=False),
                json.dumps({str(k): v for k, v in (wrong_answers or {}).items()}, ensure_ascii=False) if isinstance(wrong_answers, dict) else str(wrong_answers),
            ]
        )
        self._save_workbook(workbook)

    def list_scores(self, username: Optional[str] = None) -> List[Dict[str, object]]:
        _, sheet = self._load_sheet()
        records: List[Dict[str, object]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            if username and row[0] != username:
                continue
            records.append(
                {
                    "username": row[0],
                    "student_id": row[1],
                    "student_name": row[2],
                    "score": int(row[3] or 0),
                    "total": int(row[4] or 0),
                    "exam_time": row[5],
                    "wrong_ids": [
                        int(item)
                        for item in str(row[6] or "").split(",")
                        if item.strip().isdigit()
                    ],
                    "type_stats": json.loads(row[7] or "{}"),
                    "wrong_answers": json.loads(row[8] or "{}") if len(row) > 8 and row[8] else {},
                }
            )
        return records

    def best_score(self, username: str) -> int:
        scores = [record["score"] for record in self.list_scores(username)]
        return max(scores, default=0)

    def recent_scores(self, username: str, limit: int = 5) -> List[Dict[str, object]]:
        return self.list_scores(username)[-limit:]

    def all_wrong_ids(self, username: str) -> List[int]:
        ids: List[int] = []
        for record in self.list_scores(username):
            ids.extend(record["wrong_ids"])
        return list(dict.fromkeys(ids))

    def remove_wrong_id(self, username: str, question_id: int) -> bool:
        workbook, sheet = self._load_sheet()
        modified = False
        for row_idx in range(2, sheet.max_row + 1):
            if str(sheet.cell(row_idx, 1).value) == username:
                wrong_str = str(sheet.cell(row_idx, 7).value or "")
                ids = [int(x) for x in wrong_str.split(",") if x.strip().isdigit()]
                if question_id in ids:
                    ids.remove(question_id)
                    sheet.cell(row_idx, 7).value = ",".join(str(x) for x in ids)
                    modified = True
        if modified:
            self._save_workbook(workbook)
        return modified

    def delete_record(self, username: str, exam_time: str) -> bool:
        workbook, sheet = self._load_sheet()
        for row_idx in range(2, sheet.max_row + 1):
            if (str(sheet.cell(row_idx, 1).value) == username
                    and str(sheet.cell(row_idx, 6).value) == exam_time):
                sheet.delete_rows(row_idx, 1)
                self._save_workbook(workbook)
                return True
        return False

class ExamTracker:

    def __init__(self, path: Path | None = None) -> None:
        self.path = path or (DATA_DIR / "completed_exams.json")
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self._ensure_file()

    def _ensure_file(self) -> None:
        if not self.path.exists():
            self.path.write_text("{}", encoding="utf-8")

    def _load(self) -> Dict[str, List[str]]:
        try:
            return json.loads(self.path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return {}

    def _save(self, data: Dict[str, List[str]]) -> None:
        self.path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    def is_completed(self, username: str, exam_uuid: str) -> bool:
        data = self._load()
        return exam_uuid in data.get(username, [])

    def mark_completed(self, username: str, exam_uuid: str) -> None:
        data = self._load()
        data.setdefault(username, []).append(exam_uuid)
        self._save(data)

def export_question_bank_to_bytes(questions: List[Question], password: str,
                                   duration_minutes: int = 30,
                                   distribution: dict | None = None,
                                   exam_type: str = "exam") -> bytes:
    import uuid
    from crypto_utils import encrypt_question_bank

    exam_uuid = str(uuid.uuid4())
    payload = {
        "exam_id": exam_uuid,
        "duration_minutes": duration_minutes,
        "distribution": distribution,
        "exam_type": exam_type,
        "questions": [
            {
                "id": q.id,
                "qtype": q.qtype,
                "text": q.text,
                "options": q.options,
                "answer": q.answer,
                "analysis": q.analysis,
                "score": q.score,
            }
            for q in questions
        ],
    }
    plain = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    return exam_uuid, encrypt_question_bank(plain, password)

def import_question_bank_as_teacher(data: bytes) -> tuple:
    from crypto_utils import decrypt_question_bank_as_teacher

    plain = decrypt_question_bank_as_teacher(data)
    payload = json.loads(plain.decode("utf-8"))
    questions = [
        Question(
            id=q["id"],
            qtype=q["qtype"],
            text=q["text"],
            options=q.get("options", []),
            answer=q["answer"],
            analysis=q.get("analysis", ""),
            score=q.get("score", 10),
        )
        for q in payload["questions"]
    ]
    return payload["exam_id"], payload.get("duration_minutes", 30), payload.get("distribution"), payload.get("exam_type", "exam"), questions

def import_question_bank_as_student(data: bytes, password: str) -> tuple:
    from crypto_utils import decrypt_question_bank_as_student

    plain = decrypt_question_bank_as_student(data, password)
    payload = json.loads(plain.decode("utf-8"))
    questions = [
        Question(
            id=q["id"],
            qtype=q["qtype"],
            text=q["text"],
            options=q.get("options", []),
            answer=q["answer"],
            analysis=q.get("analysis", ""),
            score=q.get("score", 10),
        )
        for q in payload["questions"]
    ]
    return payload["exam_id"], payload.get("duration_minutes", 30), payload.get("distribution"), payload.get("exam_type", "exam"), questions

class SettingsManager:

    def __init__(self, path: Path = SETTINGS_FILE) -> None:
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self._ensure_file()

    def _ensure_file(self) -> None:
        if not self.path.exists():
            self.save_settings(DEFAULT_SETTINGS)

    def load_settings(self) -> dict:
        try:
            with self.path.open("r", encoding="utf-8") as file:
                data = json.load(file)

            merged = _deep_merge(DEFAULT_SETTINGS, data)
            return merged
        except (json.JSONDecodeError, OSError):
            return dict(DEFAULT_SETTINGS)

    def save_settings(self, settings: dict) -> None:
        with self.path.open("w", encoding="utf-8") as file:
            json.dump(settings, file, ensure_ascii=False, indent=2)

    def reset_defaults(self) -> dict:
        self.save_settings(DEFAULT_SETTINGS)
        return dict(DEFAULT_SETTINGS)

def _deep_merge(base: dict, override: dict) -> dict:
    result = dict(base)
    for key, value in override.items():
        if key in result and isinstance(result[key], dict) and isinstance(value, dict):
            result[key] = _deep_merge(result[key], value)
        else:
            result[key] = value
    return result

def process_avatar_file(filepath: str) -> str:
    import base64
    from io import BytesIO

    from PIL import Image

    img = Image.open(filepath)
    if img.mode in ("RGBA", "P"):
        img = img.convert("RGB")
    if img.width > 1024 or img.height > 1024:
        img.thumbnail((1024, 1024), Image.LANCZOS)
    buf = BytesIO()
    img.save(buf, format="JPEG", quality=80)
    return base64.b64encode(buf.getvalue()).decode("utf-8")

def avatar_to_tk_image(avatar_b64: str, size: int = 64) -> object:
    import base64
    from io import BytesIO

    from PIL import Image, ImageDraw

    raw = base64.b64decode(avatar_b64)
    img = Image.open(BytesIO(raw)).resize((size, size), Image.LANCZOS)

    mask = Image.new("L", (size, size), 0)
    ImageDraw.Draw(mask).ellipse((0, 0, size, size), fill=255)
    img.putalpha(mask)

    from tkinter import PhotoImage
    import tkinter as tk

    root = tk._default_root
    if root is None:
        import tkinter
        root = tkinter.Tk()
        root.withdraw()
    return PhotoImage(master=root, data=_pil_to_png_bytes(img))

def _pil_to_png_bytes(img) -> bytes:
    from io import BytesIO
    buf = BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()

def save_avatar(student_id: str, avatar_b64: str) -> None:
    data = _load_avatars()
    data[student_id] = avatar_b64
    AVATARS_FILE.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")

def load_avatar(student_id: str) -> str:
    return _load_avatars().get(student_id, "")

def _load_avatars() -> dict:
    if not AVATARS_FILE.exists():
        return {}
    try:
        return json.loads(AVATARS_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}